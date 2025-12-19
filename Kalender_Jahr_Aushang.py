# -*- coding: utf-8 -*-
import calendar
import locale
import pandas as pd
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.worksheet.page import PageMargins
from datetime import datetime
import sys

# Force UTF-8 encoding
if sys.version_info[0] >= 3:
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')


# Set locale to German
locale.setlocale(locale.LC_TIME, 'de_DE')

# Define year
year = 2026

# List of additional columns
columns_list = ["A", "B", "C", "D", "Alle"]

# Function to calculate Easter Sunday
def calculate_easter_sunday(year):
    a = year % 19
    b = year // 100
    c = year % 100
    d = b // 4
    e = b % 4
    f = (b + 8) // 25
    g = (b - f + 1) // 3
    h = (19 * a + b - d - g + 15) % 30
    i = c // 4
    k = c % 4
    l = (32 + 2 * e + 2 * i - h - k) % 7
    m = (a + 11 * h + 22 * l) // 451
    month = (h + l - 7 * m + 114) // 31
    day = ((h + l - 7 * m + 114) % 31) + 1
    return datetime(year, month, day)

# Function to get holidays in Bavaria
def get_bavarian_holidays(year):
    easter_sunday = calculate_easter_sunday(year)
    holidays = {
        "NJ": datetime(year, 1, 1),
        "HDK": datetime(year, 1, 6),
        "Karfr": easter_sunday - pd.Timedelta(days=2),
        "Ostrn": easter_sunday + pd.Timedelta(days=1),
        "TdArb": datetime(year, 5, 1),
        "Himfa": easter_sunday + pd.Timedelta(days=39),
        "Pfings": easter_sunday + pd.Timedelta(days=50),
        "Fron": easter_sunday + pd.Timedelta(days=60),
        "MarHim": datetime(year, 8, 15),
        "TdDE": datetime(year, 10, 3),
        "Allerh": datetime(year, 11, 1),
        "W1": datetime(year, 12, 25),
        "W2": datetime(year, 12, 26)
    }
    return holidays

# Function to calculate the ISO week number
def iso_week_number(year, month, day):
    return pd.Timestamp(year, month, day).isocalendar()[1]

holidays = get_bavarian_holidays(year)
# Create a set of holiday dates for faster lookup
holiday_dates = set(holidays.values())

# Save Excel file
filename = f"Kalender_{year}.xlsx"
with pd.ExcelWriter(filename, engine='openpyxl') as writer:
    for month in range(1, 13):
        # Generate calendar for the month
        cal = calendar.Calendar(firstweekday=0)
        days = cal.itermonthdays2(year, month)

        # Prepare data for the DataFrame
        data = []
        for day, weekday in days:
            if day != 0:  # Ignore days outside the month
                weekday_name = calendar.day_abbr[weekday]
                week_number = iso_week_number(year, month, day)
                date = datetime(year, month, day)
                
                if weekday_name == "Mo":
                    data.append([f"{day} {weekday_name} [{week_number}]"])
                else:
                    data.append([f"{day} {weekday_name}"])
                
                # Add holiday names
                for holiday_name, holiday_date in holidays.items():
                    if date == holiday_date:
                        data[-1][0] += f" {holiday_name}"

        # Create DataFrame
        month_name = calendar.month_name[month]
        df = pd.DataFrame(data, columns=[month_name])
        for column in columns_list:
            df[column] = ""

        # Create worksheet name
        worksheet_name = f"{calendar.month_abbr[month]}-{str(year)[-2:]}"

        # Write DataFrame to the worksheet
        df.to_excel(writer, index=False, sheet_name=worksheet_name)

        # Adjust column width to fit within an A4 page width
        worksheet = writer.sheets[worksheet_name]
        
        total_width = 95  # A4 width in char        
        date_column_width = len("22 Mo [8]...")
        remaining_width = total_width - date_column_width
        column_width = remaining_width / len(columns_list)

        worksheet.column_dimensions['A'].width = date_column_width
        for col_num in range(2, len(columns_list) + 2):
            worksheet.column_dimensions[chr(64 + col_num)].width = column_width

        # Mark cells with gray background for Saturdays, Sundays, and holidays
        fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        for row in range(2, len(df) + 2):
            date_cell = worksheet.cell(row=row, column=1)
            cell_value = date_cell.value
            
            # Check if weekend
            is_weekend = "Sa" in cell_value or "So" in cell_value
            
            # Check if holiday (extract day number from cell)
            day_num = int(cell_value.split()[0])
            current_date = datetime(year, month, day_num)
            is_holiday = current_date in holiday_dates
            
            if is_weekend or is_holiday:
                date_cell.fill = fill

        # Set margins: top to 5 cm, right, left, and bottom to 1 cm
        worksheet.page_margins = PageMargins(top=4 / 2.54, right=1 / 2.54, left=1 / 2.54, bottom=1 / 2.54)  # Convert cm to inches

        # Adjust row height to fit a 31-day month within margins on an A4 page
        page_height_cm = 29.7 - (5 + 1)  # A4 height minus top and bottom margins in cm
        row_height_cm = page_height_cm / 31  # Height per row in cm for a 31-day month
        row_height_inch = row_height_cm / 2.54  # Convert cm to inches

        for row in range(2, len(df) + 2):  # Excel rows start at 1, and the first row is the header
            worksheet.row_dimensions[row].height = row_height_inch * 72  # Convert inches to points (Excel uses points)
        
        # Add gray border around all cells
        border_style = Border(left=Side(style='thin', color='D3D3D3'),
                              right=Side(style='thin', color='D3D3D3'),
                              top=Side(style='thin', color='D3D3D3'),
                              bottom=Side(style='thin', color='D3D3D3'))
        
        for row in range(1, len(df) + 2):  # Include header row
            for col_num in range(1, len(columns_list) + 2): 
                cell = worksheet.cell(row=row, column=col_num)
                cell.border = border_style

    # Ensure at least one sheet is visible
    writer.book.active = 0

print(f"Die Excel-Datei '{filename}' wurde erfolgreich erstellt.")
